"""
Генератор отчетов в различных форматах.

Этот модуль предоставляет функции для создания отчетов
в форматах Excel и PDF на основе данных из базы данных.
"""

import os
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
    REPORTS_AVAILABLE = True
except ImportError:
    REPORTS_AVAILABLE = False
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from database.models import RouteProgress, Route, User, Delivery


async def generate_excel_report(
    session: AsyncSession,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    report_type: str = "general"
) -> str:
    """
    Генерирует отчет в формате Excel.
    
    Args:
        session: Сессия SQLAlchemy
        start_date: Начальная дата периода
        end_date: Конечная дата периода
        report_type: Тип отчета ("general", "couriers", "routes")
    
    Returns:
        Путь к сгенерированному файлу
    """
    if not REPORTS_AVAILABLE:
        raise ImportError("Библиотеки для генерации отчетов не установлены")
    # Создаем рабочую книгу
    wb = Workbook()
    
    # Настраиваем стили
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    if report_type == "general":
        # Общая статистика
        ws = wb.active
        ws.title = "Общая статистика"
        
        # Заголовки
        headers = [
            "Дата",
            "Всего маршрутов",
            "Всего коробок",
            "Среднее время (мин)",
            "Активные курьеры"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        
        # Получаем данные
        query = (
            select(
                func.date(RouteProgress.visited_at).label('date'),
                func.count(RouteProgress.id).label('total_routes'),
                func.sum(RouteProgress.containers_count).label('total_containers'),
                func.count(func.distinct(RouteProgress.user_id)).label('active_couriers')
            )
            .group_by(func.date(RouteProgress.visited_at))
            .order_by(func.date(RouteProgress.visited_at).desc())
        )
        
        if start_date:
            query = query.filter(RouteProgress.visited_at >= start_date)
        if end_date:
            query = query.filter(RouteProgress.visited_at <= end_date)
        
        result = await session.execute(query)
        rows = result.all()
        
        # Заполняем данные
        for row_idx, row in enumerate(rows, 2):
            ws.cell(row=row_idx, column=1, value=row.date.strftime("%d.%m.%Y"))
            ws.cell(row=row_idx, column=2, value=row.total_routes)
            ws.cell(row=row_idx, column=3, value=row.total_containers)
            ws.cell(row=row_idx, column=4, value=0)  # TODO: добавить расчет среднего времени
            ws.cell(row=row_idx, column=5, value=row.active_couriers)
            
            # Применяем границы
            for col in range(1, 6):
                ws.cell(row=row_idx, column=col).border = border
        
        # Автоматическая ширина колонок
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
    
    elif report_type == "couriers":
        # Статистика курьеров
        ws = wb.active
        ws.title = "Статистика курьеров"
        
        # Заголовки
        headers = [
            "Курьер",
            "Всего маршрутов",
            "Всего коробок",
            "Среднее коробок/маршрут",
            "Среднее время/маршрут"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        
        # Получаем данные
        query = (
            select(
                User,
                func.count(RouteProgress.id).label('total_routes'),
                func.sum(RouteProgress.containers_count).label('total_containers')
            )
            .join(RouteProgress)
            .group_by(User.telegram_id)
            .order_by(func.sum(RouteProgress.containers_count).desc())
        )
        
        if start_date:
            query = query.filter(RouteProgress.visited_at >= start_date)
        if end_date:
            query = query.filter(RouteProgress.visited_at <= end_date)
        
        result = await session.execute(query)
        rows = result.all()
        
        # Заполняем данные
        for row_idx, (user, total_routes, total_containers) in enumerate(rows, 2):
            ws.cell(row=row_idx, column=1, value=user.username or str(user.telegram_id))
            ws.cell(row=row_idx, column=2, value=total_routes)
            ws.cell(row=row_idx, column=3, value=total_containers)
            ws.cell(row=row_idx, column=4, value=round(total_containers/total_routes if total_routes else 0, 2))
            ws.cell(row=row_idx, column=5, value=0)  # TODO: добавить расчет среднего времени
            
            # Применяем границы
            for col in range(1, 6):
                ws.cell(row=row_idx, column=col).border = border
        
        # Автоматическая ширина колонок
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
    
    # Сохраняем файл
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"report_{report_type}_{timestamp}.xlsx"
    filepath = os.path.join("reports", filename)
    
    # Создаем директорию если её нет
    os.makedirs("reports", exist_ok=True)
    
    wb.save(filepath)
    return filepath


async def generate_pdf_report(
    session: AsyncSession,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    report_type: str = "general"
) -> str:
    """
    Генерирует отчет в формате PDF.
    
    Args:
        session: Сессия SQLAlchemy
        start_date: Начальная дата периода
        end_date: Конечная дата периода
        report_type: Тип отчета ("general", "couriers", "routes")
    
    Returns:
        Путь к сгенерированному файлу
    """
    if not REPORTS_AVAILABLE:
        raise ImportError("Библиотеки для генерации отчетов не установлены")
    # Создаем директорию если её нет
    os.makedirs("reports", exist_ok=True)
    
    # Создаем имя файла
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"report_{report_type}_{timestamp}.pdf"
    filepath = os.path.join("reports", filename)
    
    # Создаем PDF документ
    doc = SimpleDocTemplate(
        filepath,
        pagesize=A4,
        rightMargin=72,
        leftMargin=72,
        topMargin=72,
        bottomMargin=72
    )
    
    # Получаем стили
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    normal_style = styles['Normal']
    
    # Создаем элементы документа
    elements = []
    
    if report_type == "general":
        # Заголовок
        elements.append(Paragraph("Общая статистика маршрутов", title_style))
        elements.append(Spacer(1, 12))
        
        # Добавляем период
        period_text = "Период: "
        if start_date and end_date:
            period_text += f"с {start_date.strftime('%d.%m.%Y')} по {end_date.strftime('%d.%m.%Y')}"
        elif start_date:
            period_text += f"с {start_date.strftime('%d.%m.%Y')}"
        elif end_date:
            period_text += f"по {end_date.strftime('%d.%m.%Y')}"
        else:
            period_text += "весь период"
        
        elements.append(Paragraph(period_text, normal_style))
        elements.append(Spacer(1, 12))
        
        # Получаем данные
        query = (
            select(
                func.date(RouteProgress.visited_at).label('date'),
                func.count(RouteProgress.id).label('total_routes'),
                func.sum(RouteProgress.containers_count).label('total_containers'),
                func.count(func.distinct(RouteProgress.user_id)).label('active_couriers')
            )
            .group_by(func.date(RouteProgress.visited_at))
            .order_by(func.date(RouteProgress.visited_at).desc())
        )
        
        if start_date:
            query = query.filter(RouteProgress.visited_at >= start_date)
        if end_date:
            query = query.filter(RouteProgress.visited_at <= end_date)
        
        result = await session.execute(query)
        rows = result.all()
        
        # Создаем таблицу
        table_data = [["Дата", "Маршрутов", "Коробок", "Курьеров"]]
        for row in rows:
            table_data.append([
                row.date.strftime("%d.%m.%Y"),
                str(row.total_routes),
                str(row.total_containers),
                str(row.active_couriers)
            ])
        
        # Стиль таблицы
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.blue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ])
        
        table = Table(table_data)
        table.setStyle(table_style)
        elements.append(table)
    
    elif report_type == "couriers":
        # Заголовок
        elements.append(Paragraph("Статистика курьеров", title_style))
        elements.append(Spacer(1, 12))
        
        # Получаем данные
        query = (
            select(
                User,
                func.count(RouteProgress.id).label('total_routes'),
                func.sum(RouteProgress.containers_count).label('total_containers')
            )
            .join(RouteProgress)
            .group_by(User.telegram_id)
            .order_by(func.sum(RouteProgress.containers_count).desc())
        )
        
        if start_date:
            query = query.filter(RouteProgress.visited_at >= start_date)
        if end_date:
            query = query.filter(RouteProgress.visited_at <= end_date)
        
        result = await session.execute(query)
        rows = result.all()
        
        # Создаем таблицу
        table_data = [["Курьер", "Маршрутов", "Коробок", "Среднее"]]
        for user, total_routes, total_containers in rows:
            table_data.append([
                user.username or str(user.telegram_id),
                str(total_routes),
                str(total_containers),
                f"{round(total_containers/total_routes if total_routes else 0, 2)}"
            ])
        
        # Стиль таблицы
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.blue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ])
        
        table = Table(table_data)
        table.setStyle(table_style)
        elements.append(table)
    
    # Создаем PDF
    doc.build(elements)
    return filepath
